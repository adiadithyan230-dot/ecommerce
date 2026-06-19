import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def export_csv(df):
    """
    Exports the DataFrame to CSV bytes.
    """
    buffer = io.StringIO()
    df.to_csv(buffer, index=False)
    return buffer.getvalue().encode('utf-8')

def export_excel(df, kpis=None):
    """
    Exports the DataFrame to Excel bytes with styling.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write Main Data
        df.to_excel(writer, sheet_name='Sales Data', index=False)
        
        # Write KPI summary if available
        if kpis:
            kpi_data = {
                'Metric': [
                    'Total Revenue (Rs)', 
                    'Total Orders', 
                    'Average Order Value (Rs)', 
                    'Top Region', 
                    'Top Category', 
                    'Cancellation Rate (%)',
                    'Returning Customer Rate (%)'
                ],
                'Value': [
                    round(kpis['total_revenue'], 2),
                    kpis['total_orders'],
                    round(kpis['aov'], 2),
                    kpis['top_region'],
                    kpis['top_category'],
                    round(kpis['cancellation_rate'], 2),
                    round(kpis['returning_customer_rate'], 2)
                ]
            }
            kpi_df = pd.DataFrame(kpi_data)
            kpi_df.to_excel(writer, sheet_name='Summary Insights', index=False)
            
        # Style sheet
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            # Style header row
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                # Adjust column widths
                max_len = max(len(str(worksheet.cell(row=r, column=col).value or '')) for r in range(1, worksheet.max_row + 1))
                col_letter = worksheet.cell(row=1, column=col).column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
    return output.getvalue()

def export_pdf(df, kpis):
    """
    Generates a clean PDF report using ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=20
    )
    
    h2_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#2E75B6'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = styles['Normal']
    
    story = []
    
    # Header Section
    story.append(Paragraph("Ecommerce Analytics Performance Report", title_style))
    story.append(Paragraph(f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')} - Executive Insights", subtitle_style))
    story.append(Spacer(1, 10))
    
    # KPI Section
    story.append(Paragraph("1. Executive Key Performance Indicators", h2_style))
    
    kpi_table_data = [
        [Paragraph("<b>Metric</b>", body_style), Paragraph("<b>Performance Value</b>", body_style)],
        ["Total Revenue", f"Rs {kpis['total_revenue']:,.2f}"],
        ["Total Orders", f"{kpis['total_orders']:,}"],
        ["Average Order Value (AOV)", f"Rs {kpis['aov']:,.2f}"],
        ["Top Performing Region", str(kpis['top_region'])],
        ["Top Performing Category", str(kpis['top_category'])],
        ["Order Cancellation Rate", f"{kpis['cancellation_rate']:.2f}%"],
        ["Returning Customer Rate", f"{kpis['returning_customer_rate']:.2f}%"],
    ]
    
    kpi_table = Table(kpi_table_data, colWidths=[250, 200])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E2EFDA')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#385723')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # Top Products Summary Section
    story.append(Paragraph("2. Top 10 Product Sales Preview (Latest)", h2_style))
    
    if 'product' in df.columns and 'revenue' in df.columns:
        top_prod = df.groupby('product')['revenue'].sum().reset_index()
        top_prod = top_prod.sort_values(by='revenue', ascending=False).head(10)
        
        prod_table_data = [[Paragraph("<b>Product Name</b>", body_style), Paragraph("<b>Total Revenue Generated</b>", body_style)]]
        for _, row in top_prod.iterrows():
            prod_table_data.append([row['product'], f"Rs {row['revenue']:,.2f}"])
            
        prod_table = Table(prod_table_data, colWidths=[300, 150])
        prod_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F2F2F2')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ]))
        story.append(prod_table)
    else:
        story.append(Paragraph("Product columns not available for performance analysis.", body_style))
        
    doc.build(story)
    pdf_val = buffer.getvalue()
    buffer.close()
    return pdf_val
